#!/usr/bin/env python3
"""CLI entrypoint. Rerun this each quarter after DOL publishes a new release.

Usage:
    python run_pipeline.py
    python run_pipeline.py --force-download
    python run_pipeline.py --lca-years 3 --perm-years 2
"""
from __future__ import annotations

import argparse
import logging
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from src.discover_sources import PERFORMANCE_PAGE, SourceDiscoveryError  # noqa: E402
from src.columns import ColumnResolutionError  # noqa: E402
from src.download import DownloadError  # noqa: E402
from src.pipeline import run  # noqa: E402


def report_headers(page_url: str, lca_years: int, perm_years: int) -> int:
    """Download each selected data file, print its real header row, delete it.

    This exists because the column aliases were written without ever having
    seen a real file, and one guess in particular is load-bearing: whether
    PERM_Disclosure_Data_FY2025_Q4.xlsx uses the legacy ETA-9089 layout or
    the revised one. The filename alone can't answer that. The header row
    can, and reading it costs one pass instead of a series of failed full
    runs that each surface one missing column at a time.

    Files are deleted immediately after reading so disk stays flat, and
    only the header row is loaded so memory does too.
    """
    import openpyxl
    import requests

    from src.columns import _normalize, load_alias_config
    from src.discover_sources import USER_AGENT, get_sources
    from src.download import RAW_DIR

    sources = [
        s
        for s in get_sources(page_url=page_url, lca_years=lca_years, perm_years=perm_years)
        if s.kind != "LAYOUT"
    ]
    alias_config = load_alias_config()
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    for s in sorted(sources, key=lambda x: (x.kind, x.fiscal_year or 0)):
        name = s.url.split("/")[-1].split("?")[0]
        dest = RAW_DIR / name
        print(f"\n{'=' * 78}\n{s.kind}  FY{s.fiscal_year}  {name}\n{'=' * 78}")

        try:
            with requests.get(
                s.url, timeout=600, headers={"User-Agent": USER_AGENT}, stream=True
            ) as r:
                r.raise_for_status()
                with open(dest, "wb") as fh:
                    for chunk in r.iter_content(chunk_size=1 << 20):
                        fh.write(chunk)

            wb = openpyxl.load_workbook(dest, read_only=True)
            ws = wb[wb.sheetnames[0]]
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            wb.close()

            print(f"sheet: {wb.sheetnames[0]}   columns: {len(headers)}\n")
            for h in headers:
                print(f"  {h}")

            # Which canonical fields would resolve against this file, and
            # which would not -- the actionable part.
            aliases = alias_config.get(s.kind, {})
            present = {_normalize(h) for h in headers if h}
            missing = [
                field
                for field, cands in aliases.items()
                if not any(_normalize(c) in present for c in cands)
            ]
            if missing:
                print(f"\n  !! UNRESOLVED for kind={s.kind}: {missing}")
            else:
                print(f"\n  OK: every {s.kind} alias resolves against this file.")

        except Exception as e:  # noqa: BLE001 - report and continue to next file
            print(f"  FAILED: {type(e).__name__}: {e}")
        finally:
            if dest.exists():
                dest.unlink()

    return 0


def discover_only(page_url: str, lca_years: int, perm_years: int) -> int:
    """Report what discovery finds without downloading it.

    HEAD each candidate URL for its size so the real run can be sized in
    advance -- these files are large enough that running out of disk or
    memory mid-way is a real risk worth measuring first.
    """
    import requests

    from src.discover_sources import USER_AGENT, get_sources

    sources = get_sources(page_url=page_url, lca_years=lca_years, perm_years=perm_years)

    data_files = [s for s in sources if s.kind != "LAYOUT"]
    layouts = [s for s in sources if s.kind == "LAYOUT"]

    total_bytes = 0
    unknown_sizes = 0

    print(f"\n=== Discovery against {page_url} ===")
    print(f"Data files selected: {len(data_files)}   Record layouts: {len(layouts)}\n")

    for s in sorted(data_files, key=lambda x: (x.kind, x.fiscal_year or 0)):
        size_str = "unknown"
        try:
            r = requests.head(
                s.url, timeout=30, allow_redirects=True, headers={"User-Agent": USER_AGENT}
            )
            length = r.headers.get("Content-Length")
            if length:
                size = int(length)
                total_bytes += size
                size_str = f"{size / (1024 ** 2):.1f} MB"
            else:
                unknown_sizes += 1
        except requests.exceptions.RequestException as e:
            size_str = f"HEAD failed: {e}"
            unknown_sizes += 1

        fy = f"FY{s.fiscal_year}" if s.fiscal_year else "FY?"
        print(f"  [{s.kind:<13}] {fy:<7} {size_str:>14}   {s.url}")

    print("\n  Record layout documents:")
    for s in layouts:
        print(f"    {s.url}")

    print(f"\n=== Total known download size: {total_bytes / (1024 ** 3):.2f} GB ===")
    if unknown_sizes:
        print(f"    ({unknown_sizes} file(s) did not report a size)")

    kinds = {s.kind for s in data_files}
    print("\nSanity checks:")
    print(f"  LCA files found:          {sum(1 for s in data_files if s.kind == 'LCA')}")
    print(f"  PERM legacy files found:  {sum(1 for s in data_files if s.kind == 'PERM_LEGACY')}")
    print(f"  PERM revised files found: {sum(1 for s in data_files if s.kind == 'PERM_REVISED')}")
    if "PERM_REVISED" not in kinds:
        print(
            "\n  NOTE: no file was classified PERM_REVISED. Either OFLC hasn't split\n"
            "  the file the way the brief describes, or the classifier's keywords\n"
            "  don't match how they labelled it. Check the URLs above by eye."
        )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--page-url", default=PERFORMANCE_PAGE)
    parser.add_argument("--lca-years", type=int, default=3)
    parser.add_argument("--perm-years", type=int, default=2)
    parser.add_argument(
        "--force-download",
        action="store_true",
        help="Re-download files even if already present in data/raw/.",
    )
    parser.add_argument(
        "--discover-only",
        action="store_true",
        help="Scrape the DOL page, report which files would be pulled and how "
        "big they are, then stop. Downloads nothing. Use this to sanity-check "
        "discovery and size the real run before committing to it.",
    )
    parser.add_argument(
        "--report-headers",
        action="store_true",
        help="Download each selected data file, print its real column headers and "
        "which canonical fields fail to resolve, then delete it. Use this to fix "
        "config/column_aliases.yaml from facts instead of one failed run at a time.",
    )
    parser.add_argument("-v", "--verbose", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )

    if args.report_headers:
        try:
            return report_headers(
                page_url=args.page_url,
                lca_years=args.lca_years,
                perm_years=args.perm_years,
            )
        except SourceDiscoveryError as e:
            print(f"\nSOURCE DISCOVERY FAILED: {e}", file=sys.stderr)
            return 2

    if args.discover_only:
        try:
            return discover_only(
                page_url=args.page_url,
                lca_years=args.lca_years,
                perm_years=args.perm_years,
            )
        except SourceDiscoveryError as e:
            print(f"\nSOURCE DISCOVERY FAILED: {e}", file=sys.stderr)
            return 2

    try:
        run(
            page_url=args.page_url,
            lca_years=args.lca_years,
            perm_years=args.perm_years,
            force_download=args.force_download,
        )
    except SourceDiscoveryError as e:
        print(f"\nSOURCE DISCOVERY FAILED: {e}", file=sys.stderr)
        return 2
    except ColumnResolutionError as e:
        print(f"\nCOLUMN MAPPING FAILED: {e}", file=sys.stderr)
        return 3
    except DownloadError as e:
        print(f"\nDOWNLOAD FAILED: {e}", file=sys.stderr)
        return 4

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
